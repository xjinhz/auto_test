# --*-- encoding:utf-8 --*--
import os
import requests
import subprocess
import to_html
from log import logging
from datetime import date

try:
    from openpyxl import load_workbook
    from lxml import etree
except ImportError as e:
    subprocess.call("pip install openpyxl", shell=True)
    subprocess.call("pip install lxml", shell=True)
    from openpyxl import load_workbook
    from lxml import etree


def check_version(response):
    check_version = None
    # print(response.content)
    if "text/html" in response.headers['Content-Type']:
        page = etree.HTML((response.content.decode("utf-8")).lower())

        infos = page.xpath("//h3")

        for info in infos:
            # 原则上均能获取到版本号
            if "version:" in info.text:
                check_version = info.text
    elif "application/json" in response.headers['Content-Type']:
        check_version = eval(str(response.content, encoding="utf-8"))['version']

    # print(check_version)

    return check_version


def history_data_append(wb):
    """
    当前针对Excel表的固定单元格进行填写，要求Excel表格数据格式与模板保持一致
    :param wb:
    :return:
    """
    # wb = load_workbook("./testcase/前海API测试用例.xlsx", data_only=True)
    sheet = wb['Analysis']
    # max_row 为当前表格最大行号
    row_num = sheet.max_row + 1

    # A7~H7分别代表版本号、当前版本case总数，当前版本整体通过率、当前版本high优先级case通过率
    # Issue总数、High优先级issue总数、Medium优先级issue总数 、low优先级issue总数
    logging.info("开始追加历史版本数据")
    sheet["A%d" % row_num].value = sheet["A7"].value
    sheet["B%d" % row_num].value = sheet["B7"].value
    sheet["C%d" % row_num].value = sheet["C7"].value
    sheet["D%d" % row_num].value = sheet["D7"].value
    sheet["E%d" % row_num].value = sheet["E7"].value
    sheet["F%d" % row_num].value = sheet["F7"].value
    sheet["G%d" % row_num].value = sheet["G7"].value
    sheet["H%d" % row_num].value = sheet["H7"].value

    logging.info("历史版本数据追加完成")

    # wb.save("./testcase/" + case)

    # print(sheet["A7"].value)


def api_test():
    try:
        cases = os.listdir("./testcase/")
        for case in cases:
            # 表格中行数，调用api_test()时从第i行开始读取数据
            i = 3
            wb = load_workbook("./testcase/" + case, data_only=True)
            sheet = wb["TestCase"]
            checked_version = check_version(requests.get(sheet["G2"].value))
            if checked_version != sheet["M2"].value:
                logging.info("checked %s update and the %s is Testing" % (checked_version, case))
                while sheet["B%d" % i].value is not None:
                    try:
                        url = sheet["G%d" % i].value
                        params = sheet["H%d" % i].value.encode('utf-8')
                        headers = {'Content-Type': 'application/json'}

                        # 确认请求类型
                        if sheet["F%d" % i].value == "Post":
                            response = requests.post(url, data=params, headers=headers)
                        elif sheet["F%d" % i].value == "Get":
                            response = requests.get(url, data=params, headers=headers)
                        elif sheet["F%d" % i].value == "Put":
                            response = requests.put(url, data=params, headers=headers)
                        elif sheet["F%d" % i].value == "Update":
                            response = requests.update(url, data=params, headers=headers)
                        elif sheet["F%d" % i].value == "Delete":
                            response = requests.delete(url, data=params, headers=headers)
                        else:
                            logging.error(
                                case + " " + sheet["A%d" % i].value + " request method " + sheet["F%d" % i].value +
                                " not support, please contact the admin")
                            continue

                        # 将返回的状态码写入测试结果
                        sheet["K%d" % i] = response.status_code
                        # 判断返回的状态码与预期是否一致，不一致则直接fail，一致则进一步判断返回值是否包含相应的关键字
                        if response.status_code == sheet["I%d" % i].value:
                            if sheet["J%d" % i].value in response.content.decode("utf-8"):
                                sheet["L%d" % i] = "Pass"
                                sheet["M%d" % i] = response.content.decode("utf-8")
                                logging.info(case + " " + sheet["A%d" % i].value + response.content.decode("utf-8"))
                            else:
                                sheet["L%d" % i] = "Fail"
                                sheet["M%d" % i] = response.content.decode("utf-8")
                                logging.error(
                                    case + " " + sheet["A%d" % i].value + response.content.decode("utf-8"))
                        else:
                            sheet["L%d" % i] = "Fail"
                            sheet["M%d" % i] = response.content.decode("utf-8")
                            logging.error(case + " " + sheet["A%d" % i].value + response.content.decode("utf-8"))

                        i += 1
                    except Exception as e:
                        logging.error(case + " " + sheet["A%d" % i].value + "Error: " + str(e))
                        i += 1
                        continue

                sheet["M2"].value = checked_version
                logging.info("The new version is updated. ")
                logging.info("the %s test is completed. " % case)
                logging.info("------------------------------------------------------------------------------------\n\n")
                # 每个版本的测试数据追加到最后一行，功能尚未完成
                # history_data_append(wb)
                wb.save("./testcase/" + case)
                # 要求测试用例文档格式为XXX项目接口(或API）测试用例.xlsx
                # wb.save("./report/%s测试报告%s.html" % (case[:-10], date.today()))
                # subprocess.call("cp /opt/auto_run/testcase/%s /home/xjin/.jenkins/workspace/API_Test/" % case,
                #                 shell=True)
                with open("./report/%s测试报告%s.html" % (case[:-10], date.today()), 'w') as html_file:
                    html_file.write(to_html.HTML_TMPL)
                    html_file.flush()
                    html_file.close()
            else:
                logging.info("%s No new version found. " % case)
                logging.info("------------------------------------------------------------------------------------\n\n")
    except Exception as e:
        raise e


if __name__ == '__main__':
    api_test()
    # history_data_append()
    # check_version(response=requests.get("http://10.215.137.232:8001/main/api/backend/version/"))
