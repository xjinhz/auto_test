# --*-- encoding:utf-8 --*--
import subprocess
# import queue

try:
    from openpyxl import load_workbook
except ImportError as e:
    subprocess.call("pip install openpyxl", shell=True)
    from openpyxl import load_workbook

# 存储EXCEL中的priority, method, url and params. method 的值一般是high，medium, low
SHARE_DATA = []

# 测试用例的路径，需要填写完整路径（包括文件名称、格式），当前仅支持Excel文件
CASE_PATH = None

# 测试用例sheet页名称
SHEET_NAME = None


def get_api():
    wb = load_workbook("./testcase/APITestCase.xlsx", data_only=True)
    sheet = wb["TestCase"]
    max_row = sheet.max_row
    priority = sheet["E"]
    method = sheet["F"]
    url = sheet["G"]
    params = sheet["H"]
    global SHARE_DATA
    try:
        for row in range(max_row-1):
            # print(row)
            if str.lower(priority[row].value) == "high":
                data = {}
                data['method'] = method[row].value
                data['url'] = url[row].value
                data['params'] = params[row].value
                # print(data)

                # print(data)
                SHARE_DATA.insert(-1, data)

                # row += 1
        wb.close()
    except Exception as e:
        print(e)
    return SHARE_DATA


# if __name__ == "__main__":
#     get_api()
#     post_urls, get_urls, put_urls, patch_urls, delete_urls = [], [], [], [], []
#     post_params_queue, get_params_queue = queue.Queue(), queue.Queue()
#     put_params_queue, patch_params_queue, delete_params_queue\
#         = queue.Queue(), queue.Queue(), queue.Queue()
#
#     for i in range(len(SHARE_DATA)):
#         if SHARE_DATA[i]["method"] == "Post":
#             post_urls.append(SHARE_DATA[i]['url'])
#             # print(post_urls)
#             post_params_queue.put_nowait(SHARE_DATA[i]['params'])
#             # print(post_params_queue.get())
#         elif SHARE_DATA[i]["method"] == "Get":
#             get_urls.append(SHARE_DATA[i]['url'])
#             get_params_queue.put_nowait(SHARE_DATA[i]['params'])
#         elif SHARE_DATA[i]["method"] == "Patch":
#             patch_urls.append(SHARE_DATA[i]['url'])
#             patch_params_queue.put_nowait(SHARE_DATA[i]['params'])
#         elif SHARE_DATA[i]["method"] == "Put":
#             put_urls.append(SHARE_DATA[i]['url'])
#             put_params_queue.put_nowait(SHARE_DATA[i]['params'])
#         elif SHARE_DATA[i]["method"] == "Delete":
#             delete_urls.append(SHARE_DATA[i]['url'])
#             delete_params_queue.put_nowait(SHARE_DATA[i]['params'])
#         else:
#             continue
#     print(post_urls)
